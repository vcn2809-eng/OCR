import pytest
from pathlib import Path
import openpyxl
import csv
from app.excel_extraction.agent import (
    list_sheets, detect_header_row, extract_sheet, infer_column_types, extract_excel_document
)
from app.excel_extraction.exceptions import UnsupportedFormatError, SheetNotFoundError, NoHeaderFoundError

def test_list_sheets_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet"
    wb.create_sheet("Sheet2")
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    
    sheets = list_sheets(file_path)
    assert len(sheets) == 2
    assert "Sheet" in sheets
    assert "Sheet2" in sheets

def test_list_sheets_csv(tmp_path):
    file_path = tmp_path / "test.csv"
    with open(file_path, "w") as f:
        f.write("A,B\n1,2")
    
    assert list_sheets(file_path) == ['Sheet1']

def test_detect_header_row_clean():
    data = [
        ["Name", "Age", "City"],
        ["Alice", 30, "NY"],
        ["Bob", 25, "LA"]
    ]
    assert detect_header_row(data) == 0

def test_detect_header_row_with_title():
    data = [
        ["Q1 2024 Invoices", None, None],
        ["Invoice ID", "Amount", "Date"],
        [1, 100.5, "2024-01-01"]
    ]
    assert detect_header_row(data) == 1

def test_extract_sheet_simple(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "NY"])
    ws.append(["Bob", 25, "LA"])
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    
    res = extract_sheet(file_path, "Data")
    assert len(res) == 2
    assert res[0]["name"] == "Alice"
    assert res[1]["age"] == 25

def test_extract_sheet_merged_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Header1", "Header2", "Header3"])
    ws.append(["A1", "B1", "C1"])
    ws.merge_cells("A2:B2")
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    
    res = extract_sheet(file_path, "Data")
    assert res[0]["header1"] == "A1"
    assert res[0]["header2"] == "A1"

def test_infer_column_types_mixed():
    rows = [
        {"d": "2024-01-01", "n": 10, "c": "$100", "t": "Alice"},
        {"d": "2024-01-02", "n": 20, "c": "$200", "t": "Bob"},
        {"d": "2024-01-03", "n": 30, "c": "$300", "t": "Charlie"},
    ]
    types = infer_column_types(rows)
    assert types["d"] == "date"
    assert types["n"] == "number"
    assert types["c"] == "currency"
    assert types["t"] == "text"

def test_extract_excel_document(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.append(["Col1"])
    ws.append(["Val1"])
    
    ws2 = wb.create_sheet("S2")
    ws2.append(["Col2"])
    ws2.append(["Val2"])
    
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    
    res = extract_excel_document("doc1", file_path)
    assert "S1" in res
    assert "S2" in res
    assert res["S1"][0]["col1"] == "Val1"
    assert res["S2"][0]["col2"] == "Val2"

def test_unsupported_format_error(tmp_path):
    file_path = tmp_path / "test.txt"
    file_path.touch()
    with pytest.raises(UnsupportedFormatError):
        list_sheets(file_path)

def test_detect_header_empty_sheet():
    with pytest.raises(NoHeaderFoundError):
        detect_header_row([])
