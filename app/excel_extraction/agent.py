"""
Excel Extraction Agent — extracts structured row/column data from .xlsx, .xls, and .csv
files, handling merged cells, title rows above the real header, and formula-only cells.
"""

import csv
import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from dateutil import parser as dateutil_parser

from app.config import settings
from app.excel_extraction.exceptions import (
    ExcelExtractionError, SheetNotFoundError, NoHeaderFoundError, UnsupportedFormatError
)

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

logger = logging.getLogger(__name__)

def list_sheets(file_path: Path) -> list[str]:
    """List all sheet names in the given spreadsheet file."""
    ext = file_path.suffix.lower()
    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(file_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        elif ext == '.xls':
            if not HAS_XLRD:
                raise UnsupportedFormatError("xlrd is required for .xls files")
            wb = xlrd.open_workbook(str(file_path))
            return wb.sheet_names()
        elif ext == '.csv':
            return ['Sheet1']
        else:
            raise UnsupportedFormatError(f"Unsupported extension: {ext}")
    except UnsupportedFormatError:
        raise
    except Exception as e:
        raise ExcelExtractionError(f"Failed to open file {file_path}: {e}")

def detect_header_row(sheet_data: list[list]) -> int:
    """Detect the index of the header row in a 2D sheet data grid.

    A header row is the first row where:
    - At least 2 non-empty cells exist (single-cell title rows are skipped), AND
    - >= 50% of non-empty cells are non-numeric text strings of length <= 50.
    Falls back to 0 if no such row is found.
    """
    if not sheet_data:
        raise NoHeaderFoundError("Sheet is empty.")
    
    for idx, row in enumerate(sheet_data[:15]):
        non_empty_count = 0
        valid_count = 0
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                non_empty_count += 1
                val_str = str(cell).strip()
                if len(val_str) <= 50:
                    try:
                        float(val_str)
                    except ValueError:
                        valid_count += 1
        
        # Require >= 2 qualifying cells to distinguish headers from title rows
        if non_empty_count >= 2 and valid_count >= 2 and (valid_count / non_empty_count) >= 0.5:
            return idx
            
    return 0


def resolve_merged_cells(file_path: Path, sheet_name: str) -> list[list]:
    """Resolve merged cells by duplicating the top-left value into all merged cells."""
    ext = file_path.suffix.lower()
    if ext in ('.xls', '.csv'):
        logger.warning(f"Merged cells resolution is not supported for {ext} files.")
        return _get_raw_rows(file_path, sheet_name)
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SheetNotFoundError(f"Sheet {sheet_name} not found.")
    
    sheet = wb[sheet_name]
    
    grid = []
    for row in sheet.iter_rows():
        grid.append([cell.value for cell in row])
        
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = grid[min_row - 1][min_col - 1]
        
        for row_idx in range(min_row - 1, max_row):
            for col_idx in range(min_col - 1, max_col):
                if row_idx < len(grid) and col_idx < len(grid[row_idx]):
                    grid[row_idx][col_idx] = top_left_val
                    
    wb.close()
    return grid

def _get_raw_rows(file_path: Path, sheet_name: str) -> list[list]:
    """Get raw 2D grid from supported files without merged cell resolution."""
    ext = file_path.suffix.lower()
    if ext == '.xlsx':
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet {sheet_name} not found.")
        sheet = wb[sheet_name]
        grid = [[cell.value for cell in row] for row in sheet.iter_rows()]
        wb.close()
        return grid
    elif ext == '.xls':
        if not HAS_XLRD:
            raise UnsupportedFormatError("xlrd is required for .xls files")
        wb = xlrd.open_workbook(str(file_path))
        if sheet_name not in wb.sheet_names():
            raise SheetNotFoundError(f"Sheet {sheet_name} not found.")
        sheet = wb.sheet_by_name(sheet_name)
        return [sheet.row_values(i) for i in range(sheet.nrows)]
    elif ext == '.csv':
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            return list(reader)
    else:
        raise UnsupportedFormatError(f"Unsupported extension: {ext}")

def extract_sheet(file_path: Path, sheet_name: str) -> list[dict]:
    """Extract rows from a sheet as a list of dictionaries keyed by column name."""
    ext = file_path.suffix.lower()
    if ext == '.xlsx':
        grid = resolve_merged_cells(file_path, sheet_name)
    else:
        grid = _get_raw_rows(file_path, sheet_name)
        
    if not grid:
        return []
        
    header_idx = detect_header_row(grid)
    headers = grid[header_idx]
    
    col_names = []
    for i, h in enumerate(headers):
        if h is not None and str(h).strip():
            c = str(h).strip().lower().replace(' ', '_')
            col_names.append((i, c))
            
    rows = []
    for row_idx in range(header_idx + 1, len(grid)):
        row_data = grid[row_idx]
        if not any(cell is not None and str(cell).strip() != '' for cell in row_data):
            continue
            
        row_dict = {}
        for col_idx, col_name in col_names:
            val = row_data[col_idx] if col_idx < len(row_data) else None
            # Basic formula handling logging.
            if val is None:
                logger.debug(f"Formula cell (or empty) at {sheet_name} R{row_idx+1} C{col_idx+1}")
            row_dict[col_name] = val
        rows.append(row_dict)
        
    return rows

def infer_column_types(rows: list[dict]) -> dict[str, str]:
    """Infer column data types based on heuristics."""
    if not rows:
        return {}
        
    keys = rows[0].keys()
    col_types = {}
    
    sample_size = min(20, len(rows))
    
    for key in keys:
        valid_vals = [r[key] for r in rows[:sample_size] if r.get(key) is not None and str(r.get(key)).strip() != '']
        if not valid_vals:
            col_types[key] = 'text'
            continue
            
        date_count = 0
        currency_count = 0
        number_count = 0
        
        for val in valid_vals:
            val_str = str(val).strip()
            
            # 1. Date — require a separator char and minimum length to avoid
            # treating plain numbers (e.g. '5') as dates via fuzzy parsing.
            _DATE_SEPARATORS = ('/', '-', '.', ' ', ',')
            _has_sep = any(sep in val_str for sep in _DATE_SEPARATORS)
            if _has_sep and len(val_str) >= 5:
                try:
                    dateutil_parser.parse(val_str, fuzzy=False)
                    date_count += 1
                except (ValueError, OverflowError, TypeError):
                    pass
                
            # 2. Currency
            if any(sym in val_str for sym in ('$', '£', '€', '¥', '₹')):
                c_str = val_str.replace('$', '').replace('£', '').replace('€', '').replace('¥', '').replace('₹', '').replace(',', '').strip()
                try:
                    float(c_str)
                    currency_count += 1
                except ValueError:
                    pass
                    
            # 3. Number
            try:
                float(val_str)
                number_count += 1
            except ValueError:
                pass
                
        total = len(valid_vals)
        if date_count / total >= 0.6:
            col_types[key] = 'date'
        elif currency_count / total >= 0.6:
            col_types[key] = 'currency'
        elif number_count / total >= 0.6:
            col_types[key] = 'number'
        else:
            col_types[key] = 'text'
            
    return col_types

def extract_excel_document(document_id: str, file_path: Path) -> dict[str, list[dict]]:
    """Extract an entire excel/csv document and return sheet data mappings."""
    try:
        sheets = list_sheets(file_path)
        logger.info(f"Extracting document {document_id} with {len(sheets)} sheets.")
        
        result = {}
        for sheet in sheets:
            result[sheet] = extract_sheet(file_path, sheet)
            
        return result
    except ExcelExtractionError:
        raise
    except Exception as e:
        raise ExcelExtractionError(f"Failed to extract document: {e}")
