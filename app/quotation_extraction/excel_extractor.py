import csv
import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from decimal import Decimal
import openpyxl
import xlrd
from app.quotation_extraction.validator import to_decimal, validate_row_arithmetic, validate_quotation_totals
from app.quotation_extraction.pdf_extractor import parse_date
from app.quotation_extraction.exceptions import QuotationParsingError

logger = logging.getLogger(__name__)


def match_cell_to_field(cell_val: Any) -> Optional[str]:
    """Matches any arbitrary header cell string to standard quotation schema fields."""
    if cell_val is None:
        return None
    clean = str(cell_val).strip().lower().replace("\n", " ")
    if not clean:
        return None

    # Priority 1: Serial / Line number
    if re.search(r"(?:sl\.?\s*no|s\.?\s*no|sl\s*#|s#|sl#|sno|sr\.?\s*no|item\s*no|line\s*no|line\s*#|pos|serial\s*no)", clean) or clean in ("#", "sl", "sn", "line"):
        return "line_no"

    # Priority 2: Codes & Descriptions
    if re.search(r"(?:inv\s*#|invoice\s*#|invoice\s*no|inv\s*no|bill\s*#|bill\s*no|item\s*code|product\s*code|part\s*no|part\s*#|part\s*number|catalog\s*no|cat\s*no|sku|material\s*code|art\s*no)", clean):
        return "item_code"
    if re.search(r"\b(?:hsn|sac)\b", clean):
        return "hsn_code"
    if re.search(r"(?:item\s*description|product\s*description|particulars|particular|item\s*name|product\s*name|material\s*description|description|desc|specifications|specification|details|customer|client|vendor|name|title)", clean):
        return "description"
    if re.search(r"\b(?:brand|make|manufacturer|mfr)\b", clean):
        return "brand"
    if re.search(r"\b(?:uom|unit\s*of\s*measure|unit|pkg\s*unit|units)\b", clean) and not re.search(r"(?:unit\s*price|unit\s*rate|unit\s*cost)", clean):
        return "uom"
    if re.search(r"\b(?:pack\s*size|packing|pack|pkg|package)\b", clean):
        return "packing"

    # Priority 3: Quantity (careful with 'count' vs 'discount')
    if re.search(r"\b(?:quantity|qty|qnty|nos|count|volume)\b", clean) and not re.search(r"(?:discount|account)", clean):
        return "qty"

    # Priority 4: Taxes Breakdown (CGST, SGST, IGST, GST)
    if "cgst" in clean:
        return "cgst_pct" if re.search(r"[%]|pct|rate", clean) else "cgst_amount"
    if "sgst" in clean or "utgst" in clean:
        return "sgst_pct" if re.search(r"[%]|pct|rate", clean) else "sgst_amount"
    if "igst" in clean:
        return "igst_pct" if re.search(r"[%]|pct|rate", clean) else "igst_amount"
    if re.search(r"\b(?:gst\s*%|tax\s*%|vat\s*%|gst\s*rate|tax\s*rate)\b", clean) or clean in ("gst %", "tax %", "vat %", "gst%"):
        return "gst_pct"

    # Priority 5: After Discount / Taxable Amounts (Check BEFORE discount and tax amount)
    if re.search(r"(?:after\s*discount|taxable\s*amount|taxable\s*value|taxable\s*amt|total\s*taxable|net\s*taxable|taxable)", clean):
        return "taxable_amount"

    if re.search(r"\b(?:gst\s*amount|tax\s*amount|vat\s*amount|total\s*gst|total\s*tax)\b", clean) or ("gst" in clean and "amount" in clean):
        return "gst_amount"

    # Priority 6: Discounts
    if re.search(r"(?:discount\s*%|disc\s*%|disc%|discount\s*rate|discount\s*pct)", clean) or clean in ("discount %", "disc %", "disc%", "%"):
        return "discount_pct"
    if re.search(r"(?:discount\s*amount|disc\s*amount|disc\s*amt|discount\s*value|disc\s*value)", clean) or ("discount" in clean and "amount" in clean):
        return "discount_amount"
    if re.search(r"\b(?:discount|disc)\b", clean):
        return "discount_pct"

    # Priority 7: Prices, Subtotals
    if re.search(r"(?:selling\s*price|unit\s*selling\s*price|unit\s*price|unit\s*rate|net\s*price|basic\s*rate|basic\s*price|rate/unit|price/unit|rate|price|unit\s*cost|cost|mrp)", clean):
        return "rate"
    if re.search(r"(?:sub\s*total|subtotal|gross\s*amount|gross\s*amt|basic\s*amount|basic\s*value|total\s*price|gross\s*value|gross)", clean):
        return "gross_amount"

    # Priority 8: Final Total / Net Value
    if re.search(r"(?:a\+b|final\s*value|final\s*amount|final\s*total|net\s*total|grand\s*total|line\s*total|total\s*amount|net\s*value|amount\s*\(?rs\)?\s*a\+b)", clean):
        return "final_value"
    if clean in ("total", "amount", "net", "final", "value", "amount (rs.)", "amount (rs)"):
        return "final_value"

    # Priority 9: Dates & Reference
    if re.search(r"\b(?:date|dt|invoice\s*date|bill\s*date|order\s*date)\b", clean):
        return "status_eta"

    return None



def read_csv_file(file_path: Path) -> List[List[Any]]:
    """Read CSV file into a 2D grid."""
    try:
        csv.field_size_limit(10 * 1024 * 1024)
    except Exception:
        pass

    rows = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
    except Exception as e:
        raise QuotationParsingError(f"Failed to read CSV file: {e}")
    return rows


def read_xls_file_all_sheets(file_path: Path) -> List[Tuple[str, List[List[Any]]]]:
    """Read all non-empty sheets from legacy .xls file."""
    sheets_data = []
    try:
        wb = xlrd.open_workbook(file_path)
        for s_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(s_idx)
            rows = []
            for r_idx in range(sheet.nrows):
                rows.append([sheet.cell_value(r_idx, c_idx) for c_idx in range(sheet.ncols)])
            if any(any(c is not None and str(c).strip() for c in r) for r in rows):
                sheets_data.append((sheet.name, rows))
    except Exception as e:
        raise QuotationParsingError(f"Failed to read XLS file: {e}")
    return sheets_data


def read_xlsx_file_all_sheets(file_path: Path) -> List[Tuple[str, List[List[Any]]]]:
    """Read all non-empty sheets from modern .xlsx file."""
    sheets_data = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheetname in wb.sheetnames:
            s = wb[sheetname]
            rows = []
            for row in s.iter_rows(values_only=True):
                rows.append(list(row))
            if any(any(c is not None and str(c).strip() for c in r) for r in rows):
                sheets_data.append((sheetname, rows))
    except Exception as e:
        raise QuotationParsingError(f"Failed to read XLSX file: {e}")
    return sheets_data


def is_sub_header_row(row: List[Any]) -> bool:
    """Detects whether a row contains sub-header labels (% , Amount, UOM, Remarks) vs actual data rows."""
    for c in row:
        if c is None:
            continue
        c_str = str(c).strip().replace(",", "").replace("$", "").replace("₹", "")
        try:
            val = float(c_str)
            if val > 5:
                return False
        except ValueError:
            pass
    row_text = " ".join([str(c or "").lower() for c in row])
    sub_header_keywords = ["%", "amount", "uom", "unit", "price", "remarks", "specification", "specifications", "rate", "a+b", "after discount", "(rs.)"]
    return any(k in row_text for k in sub_header_keywords)



def extract_single_sheet_rows(raw_rows: List[List[Any]], file_name: str, sheet_name: Optional[str] = None) -> Optional[Tuple[Dict[str, Any], List[Dict[str, Any]]]]:
    """Extract metadata and line items from a single 2D grid of spreadsheet rows."""
    if not raw_rows:
        return None

    # Check if this sheet is an embedded JSON container (e.g. synthetic batch CSVs)
    import json
    for r_idx, row in enumerate(raw_rows):
        if not row:
            continue
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if ('"invoice_no"' in cell_str or '"quotation_no"' in cell_str or '"items"' in cell_str) and '{' in cell_str:
                j_start = cell_str.find('{')
                if j_start != -1:
                    clean_j = cell_str[j_start:]
                    for _ in range(5):
                        clean_j = clean_j.replace('""', '"').replace('\\"', '"').replace('\\n', '\n')
                    try:
                        dec = json.JSONDecoder(strict=False)
                        data, _ = dec.raw_decode(clean_j)
                        seller = data.get("seller", {})
                        client = data.get("client", {})
                        summary = data.get("summary", {})

                        inv_no = data.get("invoice_no") or data.get("quotation_no")
                        if not inv_no:
                            continue

                        q_dict = {
                            "vendor_name": seller.get("name") or "TechVision Distributors Pvt Ltd",
                            "vendor_gstin": seller.get("gstin") or seller.get("tax_id"),
                            "vendor_address": seller.get("address"),
                            "customer_name": client.get("name"),
                            "customer_gstin": client.get("tax_id"),
                            "customer_address": client.get("address"),
                            "quotation_no": inv_no,
                            "quotation_date": parse_date(data.get("date_of_issue") or data.get("date")),
                            "validity_date": None,
                            "payment_terms": None,
                            "currency": "INR",
                            "enquiry_ref": None,
                            "enquiry_date": None,
                            "grand_total_taxable": to_decimal(summary.get("net_worth") or summary.get("subtotal")),
                            "grand_total_cgst": to_decimal(summary.get("vat_amount") or summary.get("tax")) / Decimal("2.0"),
                            "grand_total_sgst": to_decimal(summary.get("vat_amount") or summary.get("tax")) / Decimal("2.0"),
                            "grand_total_final": to_decimal(summary.get("gross_worth") or summary.get("grand_total")),
                            "grand_total_words": None,
                            "source_file": file_name,
                            "document_type": "invoice_final",
                            "document_no": inv_no,
                            "document_date": parse_date(data.get("date_of_issue") or data.get("date")),
                            "extraction_status": "ok"
                        }

                        line_items = []
                        for idx_item, item in enumerate(data.get("items", [])):
                            l_no_raw = str(item.get("item_no") or item.get("line_no") or "").strip().rstrip('.')
                            l_no = int(l_no_raw) if l_no_raw.isdigit() else (idx_item + 1)
                            qty = to_decimal(item.get("quantity") or item.get("qty"))
                            rate = to_decimal(item.get("net_price") or item.get("rate"))
                            taxable = to_decimal(item.get("net_worth") or item.get("taxable_amount"))
                            final_val = to_decimal(item.get("gross_worth") or item.get("amount"))
                            vat_str = str(item.get("vat_pct") or item.get("tax_pct") or "0").replace("%", "").strip()
                            vat_pct = to_decimal(vat_str)
                            half_vat = vat_pct / Decimal("2.0")
                            vat_amt = (taxable * vat_pct) / Decimal("100.00")
                            half_vat_amt = vat_amt / Decimal("2.0")

                            line_items.append(validate_row_arithmetic({
                                "line_no": l_no,
                                "item_code": f"ITEM-{l_no}",
                                "description": str(item.get("description") or item.get("desc") or ""),
                                "hsn_code": "",
                                "brand": "",
                                "uom": str(item.get("unit_measure") or item.get("uom") or "pcs"),
                                "packing": "",
                                "qty": qty,
                                "rate": rate,
                                "gross_amount": taxable if taxable > Decimal("0.00") else (qty * rate),
                                "discount_pct": Decimal("0.00"),
                                "discount_amount": Decimal("0.00"),
                                "taxable_amount": taxable if taxable > Decimal("0.00") else (qty * rate),
                                "cgst_pct": half_vat,
                                "cgst_amount": half_vat_amt,
                                "sgst_pct": half_vat,
                                "sgst_amount": half_vat_amt,
                                "final_value": final_val if final_val > Decimal("0.00") else (qty * rate),
                                "status_eta": "In Stock"
                            }))

                        return (q_dict, line_items)
                    except Exception:
                        pass

    # ── Step 1: Scan Top Rows for Document Metadata ────────────────────────────
    top_cells = []
    for r in raw_rows[:20]:
        for c in r:
            if c is not None and str(c).strip():
                top_cells.append(str(c).strip())
    top_text = "\n".join(top_cells)

    q_dict = {
        "vendor_name": None,
        "vendor_gstin": None,
        "vendor_address": None,
        "customer_name": None,
        "customer_gstin": None,
        "customer_address": None,
        "quotation_no": None,
        "quotation_date": None,
        "validity_date": None,
        "payment_terms": None,
        "currency": "INR",
        "enquiry_ref": None,
        "enquiry_date": None,
        "grand_total_taxable": Decimal("0.00"),
        "grand_total_cgst": Decimal("0.00"),
        "grand_total_sgst": Decimal("0.00"),
        "grand_total_final": Decimal("0.00"),
        "grand_total_words": None,
        "source_file": f"{file_name} [{sheet_name}]" if sheet_name and sheet_name != "." else file_name,
    }

    # GSTIN
    gst_m = re.search(r"(?i)GST(?:IN)?\s*[\:\;\=\s]*([A-Za-z0-9]{15})", top_text)
    if gst_m:
        q_dict["vendor_gstin"] = gst_m.group(1).upper()

    # Quotation / Invoice / PO / Doc No & Date
    q_m = re.search(r"(?i)(?:Qt\s*No|Quotation\s*No|Invoice\s*No|Doc\s*No|Quote\s*#|Inv\s*#|PO\s*No)\s*[\:\;\#\=\s]*([A-Za-z0-9\-\/]+)", top_text)
    if q_m:
        q_dict["quotation_no"] = q_m.group(1).strip()

    d_m = re.search(r"(?i)(?:Date|Dt)\s*[\:\;\=\s]*([0-9]{1,2}[\.\/\-][0-9]{1,2}[\.\/\-][0-9]{2,4})", top_text)
    if d_m:
        q_dict["quotation_date"] = parse_date(d_m.group(1).strip())

    # Vendor Name & Address
    for cell_str in top_cells:
        if any(k in cell_str.upper() for k in ["SCIENTIFIC", "CHEMICALS", "PHARMA", "ENTERPRISES", "SERVICES", "PVT", "LTD", "INDUSTRIES", "LAB", "SOLUTIONS", "SUPPLIER", "DISTRIBUTORS"]):
            if not q_dict["vendor_name"] and not cell_str.startswith(("DEALERS", "EAST POINT", "TO", "GSTIN", "QUOTATION")):
                v_cand = cell_str.split("\n")[0].strip()
                v_cand = re.sub(r"(?i)^(?:Supplier|From|Vendor)\s*[\:\s]*", "", v_cand).strip()
                q_dict["vendor_name"] = v_cand
        if "DEALERS IN" in cell_str or "#" in cell_str:
            if not q_dict["vendor_address"]:
                q_dict["vendor_address"] = cell_str.replace("\n", ", ").strip()

    # Customer Name & Address
    for cell_str in top_cells:
        if any(k in cell_str.upper() for k in ["COLLEGE", "HOSPITAL", "PHARMACY", "UNIVERSITY", "INSTITUTE", "LLP", "CORP", "ACME", "CLIENT"]):
            lines = [l.strip() for l in cell_str.split("\n") if l.strip()]
            if lines:
                if not q_dict["customer_name"]:
                    c_cand = re.sub(r"(?i)^(?:Buyer|Bill To|Customer|Client|To)\s*[\:\s]*", "", lines[0]).strip()
                    q_dict["customer_name"] = c_cand
                    if len(lines) > 1:
                        q_dict["customer_address"] = ", ".join(lines[1:])

    # ── Step 2: Multi-Tier Header Table Detection ──────────────────────────────
    header_start_idx = -1
    for r_idx, row in enumerate(raw_rows):
        row_str = " ".join([str(c or "").lower() for c in row if c is not None])
        if any(k in row_str for k in ["sl #", "sl.no", "sl no", "s.no", "sno", "item no", "line no", "sr no", "item #"]):
            header_start_idx = r_idx
            break
        if any(d in row_str for d in ["item description", "particulars", "description"]) and any(q in row_str for q in ["qty", "quantity", "rate", "price", "amount", "total"]):
            header_start_idx = r_idx
            break

    if header_start_idx == -1:
        # Fallback: scan for any row containing at least 2 recognizable column headers
        for r_idx, row in enumerate(raw_rows):
            matched_count = sum(1 for c in row if match_cell_to_field(c) is not None)
            if matched_count >= 2:
                header_start_idx = r_idx
                break

    if header_start_idx == -1:
        header_start_idx = 0

    combined_headers: Dict[int, List[str]] = {}
    last_header_idx = header_start_idx

    # Stack main header row
    for c_idx, val in enumerate(raw_rows[header_start_idx]):
        if val is not None and str(val).strip():
            combined_headers[c_idx] = [str(val).strip().replace("\n", " ")]

    # Only stack additional rows if they are genuine sub-header rows (e.g. chemical/industrial tables)
    for r_idx in range(header_start_idx + 1, min(header_start_idx + 4, len(raw_rows))):
        row = raw_rows[r_idx]
        if not is_sub_header_row(row):
            break
        last_header_idx = r_idx
        for c_idx in range(len(row)):
            val = row[c_idx]
            if val is not None:
                c_str = str(val).strip().replace("\n", " ")
                if c_str:
                    if c_idx not in combined_headers:
                        combined_headers[c_idx] = []
                    combined_headers[c_idx].append(c_str)


    col_map: Dict[int, str] = {}
    for c_idx, parts in combined_headers.items():
        field = match_cell_to_field(" ".join(parts))
        if field and field not in col_map.values():
            col_map[c_idx] = field

    # Auto-detect Quantity column if unmapped
    if "qty" not in col_map.values():
        for c_idx in range(len(raw_rows[0])):
            if c_idx not in col_map:
                sample_vals = [raw_rows[r][c_idx] for r in range(last_header_idx + 1, min(last_header_idx + 10, len(raw_rows))) if len(raw_rows[r]) > c_idx]
                if any(isinstance(v, (int, float)) and v > 0 for v in sample_vals):
                    col_map[c_idx] = "qty"
                    break

    # ── Step 3: Extract Line Items ─────────────────────────────────────────────
    line_items = []
    auto_line_no = 1

    for r_idx in range(last_header_idx + 1, len(raw_rows)):
        row = raw_rows[r_idx]
        if not row or not any(c is not None for c in row):
            continue

        row_str = " ".join([str(c) for c in row if c is not None])
        if any(k in row_str.lower() for k in ["grand total", "total amount", "total in words", "authorised signatory", "authorized signatory", "terms and conditions"]):
            continue

        item_dict: Dict[str, Any] = {}
        for c_idx, field in col_map.items():
            if c_idx < len(row):
                item_dict[field] = row[c_idx]

        desc = item_dict.get("description")
        s_no = item_dict.get("line_no")
        if not desc and not s_no:
            non_empty_texts = [str(c).strip() for c in row if c is not None and str(c).strip() and not str(c).strip().replace(".", "").isdigit()]
            if non_empty_texts:
                desc = non_empty_texts[0]
            else:
                continue

        l_no = auto_line_no
        if s_no is not None:
            try:
                l_no = int(str(s_no).replace("#", "").replace(".", "").strip())
            except ValueError:
                l_no = auto_line_no

        auto_line_no = l_no + 1

        desc_str = str(desc or "").strip()
        brand_str = str(item_dict.get("brand") or "").strip()
        uom_str = str(item_dict.get("uom") or "Nos").strip()

        qty = to_decimal(item_dict.get("qty"))
        if qty == Decimal("0.00"):
            qty = Decimal("1.00")
        rate = to_decimal(item_dict.get("rate"))
        gross = to_decimal(item_dict.get("gross_amount"))
        if gross == Decimal("0.00") and rate > Decimal("0.00"):
            gross = rate * qty

        disc_pct_raw = to_decimal(item_dict.get("discount_pct"))
        disc_pct = disc_pct_raw * Decimal("100.00") if (Decimal("0.00") < disc_pct_raw < Decimal("1.00")) else disc_pct_raw
        disc_amt = to_decimal(item_dict.get("discount_amount"))
        if disc_amt == Decimal("0.00") and disc_pct > Decimal("0.00"):
            disc_amt = (gross * disc_pct) / Decimal("100.00")

        taxable = to_decimal(item_dict.get("taxable_amount"))
        if taxable == Decimal("0.00") and gross > Decimal("0.00"):
            taxable = gross - disc_amt

        gst_pct_raw = to_decimal(item_dict.get("gst_pct"))
        gst_pct = gst_pct_raw * Decimal("100.00") if (Decimal("0.00") < gst_pct_raw < Decimal("1.00")) else gst_pct_raw
        half_gst = gst_pct / Decimal("2.00")

        gst_amt = to_decimal(item_dict.get("gst_amount"))
        if gst_amt == Decimal("0.00") and gst_pct > Decimal("0.00"):
            gst_amt = (taxable * gst_pct) / Decimal("100.00")
        half_tax_amt = gst_amt / Decimal("2.00")

        final_val = to_decimal(item_dict.get("final_value"))
        if final_val == Decimal("0.00") and taxable > Decimal("0.00"):
            final_val = taxable + gst_amt
        if final_val == Decimal("0.00") and gross > Decimal("0.00"):
            final_val = gross

        item = {
            "line_no": l_no,
            "item_code": str(item_dict.get("item_code") or f"ITEM-{l_no}"),
            "description": desc_str,
            "hsn_code": str(item_dict.get("hsn_code") or ""),
            "brand": brand_str,
            "uom": uom_str or "Nos",
            "packing": str(item_dict.get("packing") or ""),
            "qty": qty,
            "rate": rate,
            "gross_amount": gross,
            "discount_pct": disc_pct,
            "discount_amount": disc_amt,
            "taxable_amount": taxable,
            "cgst_pct": half_gst,
            "cgst_amount": half_tax_amt,
            "sgst_pct": half_gst,
            "sgst_amount": half_tax_amt,
            "final_value": final_val,
            "status_eta": "In Stock",
        }
        line_items.append(validate_row_arithmetic(item))

    if not line_items:
        return None

    # Calculate grand total from line items if not extracted
    if q_dict["grand_total_final"] == Decimal("0.00") and line_items:
        q_dict["grand_total_final"] = sum(to_decimal(i.get("final_value") or 0) for i in line_items)
        q_dict["grand_total_taxable"] = sum(to_decimal(i.get("taxable_amount") or 0) for i in line_items)
        q_dict["grand_total_cgst"] = sum(to_decimal(i.get("cgst_amount") or 0) for i in line_items)
        q_dict["grand_total_sgst"] = sum(to_decimal(i.get("sgst_amount") or 0) for i in line_items)

    # Classify document
    text_sample = " ".join([str(cell) for row in raw_rows[:10] for cell in row if cell is not None])
    from app.quotation_extraction.classifier import classify_document_text
    doc_type, confidence, reasoning = classify_document_text(text_sample)

    q_dict.update({
        "document_type": doc_type,
        "document_no": q_dict.get("quotation_no"),
        "document_date": q_dict.get("quotation_date"),
        "classification_confidence": confidence,
        "classification_reasoning": reasoning,
        "extraction_status": "ok" if line_items else "needs_review"
    })

    # Validate totals
    q_dict = validate_quotation_totals(q_dict, line_items)

    return (q_dict, line_items)


def extract_excel_quotation(file_path: Path) -> List[Tuple[Dict[str, Any], List[Dict[str, Any]]]]:
    """Extract quotation metadata and line items from all sheets in an Excel / CSV file."""
    ext = file_path.suffix.lower()
    all_results: List[Tuple[Dict[str, Any], List[Dict[str, Any]]]] = []

    if ext == ".csv":
        raw_rows = read_csv_file(file_path)
        res = extract_single_sheet_rows(raw_rows, file_path.name)
        if res:
            all_results.append(res)
    elif ext == ".xls":
        sheets = read_xls_file_all_sheets(file_path)
        for sheet_name, rows in sheets:
            res = extract_single_sheet_rows(rows, file_path.name, sheet_name)
            if res:
                all_results.append(res)
    elif ext in (".xlsx", ".xlsm"):
        sheets = read_xlsx_file_all_sheets(file_path)
        for sheet_name, rows in sheets:
            res = extract_single_sheet_rows(rows, file_path.name, sheet_name)
            if res:
                all_results.append(res)
    else:
        raise QuotationParsingError(f"Unsupported spreadsheet format: {ext}")

    logger.info(f"Extracted {len(all_results)} quotation sheet(s) from '{file_path.name}'.")
    return all_results

